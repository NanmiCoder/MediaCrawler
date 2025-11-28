# -*- coding: utf-8 -*-
# Copyright (c) 2025 relakkes@gmail.com
#
# This file is part of MediaCrawler project.
# Repository: https://github.com/NanmiCoder/MediaCrawler/blob/main/store/excel_store_base.py
# GitHub: https://github.com/NanmiCoder
# Licensed under NON-COMMERCIAL LEARNING LICENSE 1.1
#
# 声明：本代码仅供学习和研究目的使用。使用者应遵守以下原则：
# 1. 不得用于任何商业用途。
# 2. 使用时应遵守目标平台的使用条款和robots.txt规则。
# 3. 不得进行大规模爬取或对平台造成运营干扰。
# 4. 应合理控制请求频率，避免给目标平台带来不必要的负担。
# 5. 不得用于任何非法或不当的用途。
#
# 详细许可条款请参阅项目根目录下的LICENSE文件。
# 使用本代码即表示您同意遵守上述原则和LICENSE中的所有条款。

# 声明:本代码仅供学习和研究目的使用。使用者应遵守以下原则:
# 1. 不得用于任何商业用途。
# 2. 使用时应遵守目标平台的使用条款和robots.txt规则。
# 3. 不得进行大规模爬取或对平台造成运营干扰。
# 4. 应合理控制请求频率,避免给目标平台带来不必要的负担。
# 5. 不得用于任何非法或不当的用途。
#
# 详细许可条款请参阅项目根目录下的LICENSE文件。
# 使用本代码即表示您同意遵守上述原则和LICENSE中的所有条款。

"""
Excel Store Base Implementation
Provides Excel export functionality for crawled data with formatted sheets
"""

import threading
from datetime import datetime
from typing import Dict, List, Any
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from base.base_crawler import AbstractStore
from tools import utils


class ExcelStoreBase(AbstractStore):
    """
    Base class for Excel storage implementation
    Provides formatted Excel export with multiple sheets for contents, comments, and creators
    Uses singleton pattern to maintain state across multiple store calls
    """

    # Class-level singleton management
    _instances: Dict[str, "ExcelStoreBase"] = {}
    _lock = threading.Lock()

    @classmethod
    def get_instance(cls, platform: str, crawler_type: str) -> "ExcelStoreBase":
        """
        Get or create a singleton instance for the given platform and crawler type

        Args:
            platform: Platform name (xhs, dy, ks, etc.)
            crawler_type: Type of crawler (search, detail, creator)

        Returns:
            ExcelStoreBase instance
        """
        key = f"{platform}_{crawler_type}"
        with cls._lock:
            if key not in cls._instances:
                cls._instances[key] = cls(platform, crawler_type)
            return cls._instances[key]

    @classmethod
    def flush_all(cls):
        """
        Flush all Excel store instances and save to files
        Should be called at the end of crawler execution
        """
        with cls._lock:
            for key, instance in cls._instances.items():
                try:
                    instance.flush()
                    utils.logger.info(f"[ExcelStoreBase] Flushed instance: {key}")
                except Exception as e:
                    utils.logger.error(f"[ExcelStoreBase] Error flushing {key}: {e}")
            cls._instances.clear()

    def __init__(self, platform: str, crawler_type: str = "search"):
        """
        Initialize Excel store

        Args:
            platform: Platform name (xhs, dy, ks, etc.)
            crawler_type: Type of crawler (search, detail, creator)
        """
        if not EXCEL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

        super().__init__()
        self.platform = platform
        self.crawler_type = crawler_type

        # Create data directory
        self.data_dir = Path("data") / platform
        self.data_dir.mkdir(parents=True, exist_ok=True)

        # Initialize workbook
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet

        # Create sheets
        self.contents_sheet = self.workbook.create_sheet("Contents")
        self.comments_sheet = self.workbook.create_sheet("Comments")
        self.creators_sheet = self.workbook.create_sheet("Creators")

        # Track if headers are written
        self.contents_headers_written = False
        self.comments_headers_written = False
        self.creators_headers_written = False
        self.contacts_headers_written = False
        self.dynamics_headers_written = False

        # Optional sheets for platforms that need them (e.g., Bilibili)
        self.contacts_sheet = None
        self.dynamics_sheet = None

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.filename = self.data_dir / f"{platform}_{crawler_type}_{timestamp}.xlsx"

        utils.logger.info(f"[ExcelStoreBase] Initialized Excel export to: {self.filename}")

    def _apply_header_style(self, sheet, row_num: int = 1):
        """
        Apply formatting to header row

        Args:
            sheet: Worksheet object
            row_num: Row number for headers (default: 1)
        """
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in sheet[row_num]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

    def _auto_adjust_column_width(self, sheet):
        """
        Auto-adjust column widths based on content

        Args:
            sheet: Worksheet object
        """
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (TypeError, AttributeError):
                    pass

            # Set width with min/max constraints
            adjusted_width = min(max(max_length + 2, 10), 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _write_headers(self, sheet, headers: List[str]):
        """
        Write headers to sheet

        Args:
            sheet: Worksheet object
            headers: List of header names
        """
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        self._apply_header_style(sheet)

    def _write_row(self, sheet, data: Dict[str, Any], headers: List[str]):
        """
        Write data row to sheet

        Args:
            sheet: Worksheet object
            data: Data dictionary
            headers: List of header names (defines column order)
        """
        row_num = sheet.max_row + 1

        for col_num, header in enumerate(headers, 1):
            value = data.get(header, "")

            # Handle different data types
            if isinstance(value, (list, dict)):
                value = str(value)
            elif value is None:
                value = ""

            cell = sheet.cell(row=row_num, column=col_num, value=value)

            # Apply basic formatting
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    async def store_content(self, content_item: Dict):
        """
        Store content data to Excel

        Args:
            content_item: Content data dictionary
        """
        # Define headers (customize based on platform)
        headers = list(content_item.keys())

        # Write headers if first time
        if not self.contents_headers_written:
            self._write_headers(self.contents_sheet, headers)
            self.contents_headers_written = True

        # Write data row
        self._write_row(self.contents_sheet, content_item, headers)

        # Get ID from various possible field names
        content_id = content_item.get('note_id') or content_item.get('aweme_id') or content_item.get('video_id') or content_item.get('content_id') or 'N/A'
        utils.logger.info(f"[ExcelStoreBase] Stored content to Excel: {content_id}")

    async def store_comment(self, comment_item: Dict):
        """
        Store comment data to Excel

        Args:
            comment_item: Comment data dictionary
        """
        # Define headers
        headers = list(comment_item.keys())

        # Write headers if first time
        if not self.comments_headers_written:
            self._write_headers(self.comments_sheet, headers)
            self.comments_headers_written = True

        # Write data row
        self._write_row(self.comments_sheet, comment_item, headers)

        utils.logger.info(f"[ExcelStoreBase] Stored comment to Excel: {comment_item.get('comment_id', 'N/A')}")

    async def store_creator(self, creator: Dict):
        """
        Store creator data to Excel

        Args:
            creator: Creator data dictionary
        """
        # Define headers
        headers = list(creator.keys())

        # Write headers if first time
        if not self.creators_headers_written:
            self._write_headers(self.creators_sheet, headers)
            self.creators_headers_written = True

        # Write data row
        self._write_row(self.creators_sheet, creator, headers)

        utils.logger.info(f"[ExcelStoreBase] Stored creator to Excel: {creator.get('user_id', 'N/A')}")

    async def store_contact(self, contact_item: Dict):
        """
        Store contact data to Excel (for platforms like Bilibili)

        Args:
            contact_item: Contact data dictionary
        """
        # Create contacts sheet if not exists
        if self.contacts_sheet is None:
            self.contacts_sheet = self.workbook.create_sheet("Contacts")

        # Define headers
        headers = list(contact_item.keys())

        # Write headers if first time
        if not self.contacts_headers_written:
            self._write_headers(self.contacts_sheet, headers)
            self.contacts_headers_written = True

        # Write data row
        self._write_row(self.contacts_sheet, contact_item, headers)

        utils.logger.info(f"[ExcelStoreBase] Stored contact to Excel: up_id={contact_item.get('up_id', 'N/A')}, fan_id={contact_item.get('fan_id', 'N/A')}")

    async def store_dynamic(self, dynamic_item: Dict):
        """
        Store dynamic data to Excel (for platforms like Bilibili)

        Args:
            dynamic_item: Dynamic data dictionary
        """
        # Create dynamics sheet if not exists
        if self.dynamics_sheet is None:
            self.dynamics_sheet = self.workbook.create_sheet("Dynamics")

        # Define headers
        headers = list(dynamic_item.keys())

        # Write headers if first time
        if not self.dynamics_headers_written:
            self._write_headers(self.dynamics_sheet, headers)
            self.dynamics_headers_written = True

        # Write data row
        self._write_row(self.dynamics_sheet, dynamic_item, headers)

        utils.logger.info(f"[ExcelStoreBase] Stored dynamic to Excel: {dynamic_item.get('dynamic_id', 'N/A')}")

    def flush(self):
        """
        Save workbook to file
        """
        try:
            # Auto-adjust column widths for all sheets
            self._auto_adjust_column_width(self.contents_sheet)
            self._auto_adjust_column_width(self.comments_sheet)
            self._auto_adjust_column_width(self.creators_sheet)
            if self.contacts_sheet is not None:
                self._auto_adjust_column_width(self.contacts_sheet)
            if self.dynamics_sheet is not None:
                self._auto_adjust_column_width(self.dynamics_sheet)

            # Remove empty sheets (only header row)
            if self.contents_sheet.max_row == 1:
                self.workbook.remove(self.contents_sheet)
            if self.comments_sheet.max_row == 1:
                self.workbook.remove(self.comments_sheet)
            if self.creators_sheet.max_row == 1:
                self.workbook.remove(self.creators_sheet)
            if self.contacts_sheet is not None and self.contacts_sheet.max_row == 1:
                self.workbook.remove(self.contacts_sheet)
            if self.dynamics_sheet is not None and self.dynamics_sheet.max_row == 1:
                self.workbook.remove(self.dynamics_sheet)

            # Check if there are any sheets left
            if len(self.workbook.sheetnames) == 0:
                utils.logger.info(f"[ExcelStoreBase] No data to save, skipping file creation: {self.filename}")
                return

            # Save workbook
            self.workbook.save(self.filename)
            utils.logger.info(f"[ExcelStoreBase] Excel file saved successfully: {self.filename}")

        except Exception as e:
            utils.logger.error(f"[ExcelStoreBase] Error saving Excel file: {e}")
            raise
